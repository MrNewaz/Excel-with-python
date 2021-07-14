import openpyxl

wb=openpyxl.Workbook()

wl=openpyxl.load_workbook('newaz.xlsx')

print(wl.sheetnames)